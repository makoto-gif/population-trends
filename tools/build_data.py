#!/usr/bin/env python3
"""生データ（OWID / e-Stat / 社人研）をアプリ用の JSON に変換する。

使い方:
    python3 tools/build_data.py --raw <生データ置き場>

生データの取得元は tools/fetch_raw.sh を参照。
"""
import argparse
import csv
import json
import os
from datetime import date

from jp_names import JP, REGIONS, WORLD

WORLD_START = 1800
WORLD_END = 2100
LAST_HISTORICAL = 2023  # これ以降は国連の将来推計

CONTINENT = {
    "アジア": """Afghanistan Armenia Azerbaijan Bahrain Bangladesh Bhutan Brunei Cambodia China Cyprus
        East Timor Georgia Hong Kong India Indonesia Iran Iraq Israel Japan Jordan Kazakhstan Kuwait
        Kyrgyzstan Laos Lebanon Macao Malaysia Maldives Mongolia Myanmar Nepal North Korea Oman Pakistan
        Palestine Philippines Qatar Saudi Arabia Singapore South Korea Sri Lanka Syria Taiwan Tajikistan
        Thailand Turkey Turkmenistan United Arab Emirates Uzbekistan Vietnam Yemen""",
    "ヨーロッパ": """Albania Andorra Austria Belarus Belgium Bosnia and Herzegovina Bulgaria Croatia Czechia
        Denmark Estonia Faroe Islands Finland France Germany Gibraltar Greece Guernsey Hungary Iceland
        Ireland Isle of Man Italy Jersey Kosovo Latvia Liechtenstein Lithuania Luxembourg Malta Moldova
        Monaco Montenegro Netherlands North Macedonia Norway Poland Portugal Romania Russia San Marino
        Serbia Slovakia Slovenia Spain Svalbard and Jan Mayen Sweden Switzerland Ukraine United Kingdom
        Vatican""",
    "アフリカ": """Algeria Angola Benin Botswana Burkina Faso Burundi Cameroon Cape Verde Central African Republic
        Chad Comoros Congo Cote d'Ivoire Democratic Republic of Congo Djibouti Egypt Equatorial Guinea
        Eritrea Eswatini Ethiopia Gabon Gambia Ghana Guinea Guinea-Bissau Kenya Lesotho Liberia Libya
        Madagascar Malawi Mali Mauritania Mauritius Mayotte Morocco Mozambique Namibia Niger Nigeria
        Reunion Rwanda Saint Helena Sao Tome and Principe Senegal Seychelles Sierra Leone Somalia
        South Africa South Sudan Sudan Tanzania Togo Tunisia Uganda Western Sahara Zambia Zimbabwe""",
    "北アメリカ": """Anguilla Antigua and Barbuda Aruba Bahamas Barbados Belize Bermuda
        Bonaire Sint Eustatius and Saba British Virgin Islands Canada Cayman Islands Costa Rica Cuba
        Curacao Dominica Dominican Republic El Salvador Greenland Grenada Guadeloupe Guatemala Haiti
        Honduras Jamaica Martinique Mexico Montserrat Nicaragua Panama Puerto Rico
        Saint Barthelemy Saint Kitts and Nevis Saint Lucia Saint Martin (French part)
        Saint Pierre and Miquelon Saint Vincent and the Grenadines Sint Maarten (Dutch part)
        Trinidad and Tobago Turks and Caicos Islands United States United States Virgin Islands""",
    "南アメリカ": """Argentina Bolivia Brazil Chile Colombia Ecuador Falkland Islands French Guiana Guyana
        Paraguay Peru Suriname Uruguay Venezuela""",
    "オセアニア": """American Samoa Australia Cook Islands Fiji French Polynesia Guam Kiribati Marshall Islands
        Micronesia (country) Nauru New Caledonia New Zealand Niue Northern Mariana Islands Palau
        Papua New Guinea Samoa Solomon Islands Tokelau Tonga Tuvalu Vanuatu Wallis and Futuna""",
}

# 複数語の国名があるため、CONTINENT の文字列は JP のキーと突き合わせて解決する
def build_continent_map():
    m = {}
    for cont, blob in CONTINENT.items():
        text = " ".join(blob.split())
        remaining = text
        # 長い国名から順にマッチさせる
        for name in sorted(JP, key=len, reverse=True):
            if name in remaining:
                m[name] = cont
                remaining = remaining.replace(name, " ")
        leftover = [w for w in remaining.split() if w]
        if leftover:
            print(f"  [warn] {cont}: 未解決の語 {leftover}")
    return m


def build_world(raw_dir, out_dir):
    src = os.path.join(raw_dir, "pop_proj.csv")
    rows = list(csv.DictReader(open(src, encoding="utf-8")))

    years = list(range(WORLD_START, WORLD_END + 1))
    yindex = {y: i for i, y in enumerate(years)}

    series = {}
    codes = {}
    for r in rows:
        y = int(r["year"])
        if y < WORLD_START or y > WORLD_END:
            continue
        ent = r["entity"]
        raw = r["population_historical"] or r["population_projection__projected"]
        if not raw:
            continue
        arr = series.setdefault(ent, [None] * len(years))
        # 千人単位に丸める（ファイルサイズ削減）
        arr[yindex[y]] = round(float(raw) / 1000)
        codes.setdefault(ent, r["code"])

    cont = build_continent_map()
    entities = []

    def add(name, jp, group, region=None):
        if name not in series:
            print(f"  [warn] データなし: {name}")
            return
        entities.append({
            "id": codes[name] or name,
            "name": jp,
            "en": name,
            "group": group,
            "region": region,
            "v": series[name],
        })

    for name, jp in WORLD.items():
        add(name, jp, "world")
    for name, jp in REGIONS.items():
        add(name, jp, "region")
    for name, jp in sorted(JP.items(), key=lambda kv: kv[1]):
        add(name, jp, "country", cont.get(name, "その他"))

    data = {
        "meta": {
            "startYear": WORLD_START,
            "endYear": WORLD_END,
            "lastHistoricalYear": LAST_HISTORICAL,
            "unit": "thousand",
            "builtAt": date.today().isoformat(),
            "sourceLabel": "Our World in Data（HYDE / Gapminder / 国連 WPP）",
            "sources": [
                {"period": "1800–1949", "name": "Gapminder v7",
                 "url": "https://www.gapminder.org/data/documentation/gd003/"},
                {"period": "1950–2023（実績）", "name": "国連 World Population Prospects 2024",
                 "url": "https://population.un.org/wpp/"},
                {"period": "2024–2100（推計）", "name": "国連 World Population Prospects 2024（中位推計）",
                 "url": "https://population.un.org/wpp/"},
                {"period": "全期間の統合", "name": "Our World in Data",
                 "url": "https://ourworldindata.org/grapher/population-with-projections"},
            ],
        },
        "entities": entities,
    }
    path = os.path.join(out_dir, "world.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"世界: {len(entities)} 件 → {path} ({os.path.getsize(path)/1024:.0f} KB)")


PREF_ORDER = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県", "茨城県", "栃木県", "群馬県",
    "埼玉県", "千葉県", "東京都", "神奈川県", "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県", "福岡県",
    "佐賀県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県", "長崎県",
]

PREF_REGION = {
    "北海道": "北海道・東北", "青森県": "北海道・東北", "岩手県": "北海道・東北", "宮城県": "北海道・東北",
    "秋田県": "北海道・東北", "山形県": "北海道・東北", "福島県": "北海道・東北",
    "茨城県": "関東", "栃木県": "関東", "群馬県": "関東", "埼玉県": "関東",
    "千葉県": "関東", "東京都": "関東", "神奈川県": "関東",
    "新潟県": "中部", "富山県": "中部", "石川県": "中部", "福井県": "中部", "山梨県": "中部",
    "長野県": "中部", "岐阜県": "中部", "静岡県": "中部", "愛知県": "中部",
    "三重県": "近畿", "滋賀県": "近畿", "京都府": "近畿", "大阪府": "近畿",
    "兵庫県": "近畿", "奈良県": "近畿", "和歌山県": "近畿",
    "鳥取県": "中国・四国", "島根県": "中国・四国", "岡山県": "中国・四国", "広島県": "中国・四国",
    "山口県": "中国・四国", "徳島県": "中国・四国", "香川県": "中国・四国", "愛媛県": "中国・四国",
    "高知県": "中国・四国",
    "福岡県": "九州・沖縄", "佐賀県": "九州・沖縄", "長崎県": "九州・沖縄", "熊本県": "九州・沖縄",
    "大分県": "九州・沖縄", "宮崎県": "九州・沖縄", "鹿児島県": "九州・沖縄", "沖縄県": "九州・沖縄",
}


def build_japan(raw_dir, out_dir):
    import openpyxl

    census_years = list(range(1920, 2016, 5))
    proj_years = list(range(2020, 2051, 5))
    years = census_years + proj_years
    yindex = {y: i for i, y in enumerate(years)}

    series = {}   # 名前 -> 配列（人）
    codes = {}

    # --- 国勢調査 1920–2015（総務省, e-Stat） ---
    with open(os.path.join(raw_dir, "pref.csv"), encoding="cp932") as f:
        for row in csv.reader(f):
            if len(row) < 7 or not row[0].strip().isdigit():
                continue
            code, name, year, total = row[0], row[1], row[4], row[6]
            if not year.strip().isdigit():
                continue
            y = int(year)
            if y not in yindex:
                continue
            arr = series.setdefault(name, [None] * len(years))
            total = str(total).strip().replace(",", "")
            # 未調査の年は "-" が入る（例: 1945年の沖縄県）
            arr[yindex[y]] = int(total) if total.isdigit() else None
            codes.setdefault(name, code)

    # --- 2020 実績 + 2025–2050 推計（国立社会保障・人口問題研究所） ---
    wb = openpyxl.load_workbook(os.path.join(raw_dir, "kekkahyo1.xlsx"),
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    national = [0] * len(proj_years)
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[1] != "a":      # a = 都道府県の行
            continue
        name = row[2]
        arr = series.setdefault(name, [None] * len(years))
        for i, y in enumerate(proj_years):
            val = row[4 + i]
            if val is None:
                continue
            arr[yindex[y]] = int(val)
            national[i] += int(val)
    nat = series["全国"]
    for i, y in enumerate(proj_years):
        nat[yindex[y]] = national[i]

    entities = []
    for name in ["全国"] + PREF_ORDER:
        if name not in series:
            print(f"  [warn] データなし: {name}")
            continue
        entities.append({
            "id": codes.get(name, name),
            "name": name,
            "en": name,
            "group": "national" if name == "全国" else "prefecture",
            "region": PREF_REGION.get(name, "全国"),
            "v": series[name],
        })

    data = {
        "meta": {
            "years": years,
            "lastHistoricalYear": 2020,
            "unit": "person",
            "builtAt": date.today().isoformat(),
            "sourceLabel": "総務省統計局「国勢調査」／国立社会保障・人口問題研究所「日本の地域別将来推計人口」",
            "sources": [
                {"period": "1920–2015（実績）", "name": "総務省統計局「国勢調査」時系列データ",
                 "url": "https://www.e-stat.go.jp/stat-search/database?tstat=000001011777"},
                {"period": "2020（実績）", "name": "総務省統計局「令和2年国勢調査」",
                 "url": "https://www.e-stat.go.jp/stat-search/database?tstat=000001011777"},
                {"period": "2025–2050（推計）",
                 "name": "国立社会保障・人口問題研究所「日本の地域別将来推計人口（令和5年推計）」",
                 "url": "https://www.ipss.go.jp/pp-shicyoson/j/shicyoson23/t-page.asp"},
            ],
            "notes": [
                "1945年は「人口調査」による。沖縄県は調査されていないため含まれない。",
                "1920–1970年の沖縄県は米国施政権下の統計を含む年がある。",
                "2020年の値は国勢調査の実績（社人研の推計基準年）。",
            ],
        },
        "entities": entities,
    }
    path = os.path.join(out_dir, "japan.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"日本: {len(entities)} 件 → {path} ({os.path.getsize(path)/1024:.0f} KB)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", required=True, help="生データのディレクトリ")
    ap.add_argument("--out", default=None, help="出力先（既定: ../data）")
    args = ap.parse_args()
    out = args.out or os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
    os.makedirs(out, exist_ok=True)
    build_world(args.raw, out)
    build_japan(args.raw, out)


if __name__ == "__main__":
    main()
